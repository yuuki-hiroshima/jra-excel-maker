# v0.14 (改善版：複数アプローチ併用)
import os, re, threading
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

import requests
from bs4 import BeautifulSoup
from bs4.element import Tag
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# tkcalendar は任意（ある場合はポップアップカレンダー使用）
TKCAL_OK = True
try:
    from tkcalendar import DateEntry, Calendar
except Exception:
    TKCAL_OK = False

VENUES = ["札幌","函館","福島","新潟","東京","中山","中京","京都","阪神","小倉"]
VENUE_CODE = {
    "札幌":"01","函館":"02","福島":"03","新潟":"04",
    "東京":"05","中山":"06","中京":"07","京都":"08",
    "阪神":"09","小倉":"10",
}

# ================== HTMLユーティリティ ==================
def anchor_text(cell: Tag) -> str:
    if hasattr(cell, "find"):
        a = cell.find("a")
        if a:
            t = a.get_text(strip=True)
            if t:
                return t
        return cell.get_text(strip=True)
    return str(cell).strip()

def clean_name(s: str) -> str:
    s = re.split(r"[（(]", s, maxsplit=1)[0]
    return s.strip()

def find_col_index(header_map: dict, candidates) -> int | None:
    for key in candidates:
        for h, idx in header_map.items():
            if h == key: return idx
    for key in candidates:
        for h, idx in header_map.items():
            if key in h: return idx
    return None

def find_table_and_headers(soup: BeautifulSoup):
    for t in soup.find_all("table"):
        thead = t.find("thead")
        head_cells = thead.find_all(["th","td"]) if thead else (t.find("tr").find_all(["th","td"]) if t.find("tr") else [])
        if not head_cells:
            continue
        heads_raw = [c.get_text(strip=True) for c in head_cells]
        heads_norm = [re.sub(r"\s+","", h) for h in heads_raw]
        has_horse = any("馬名" in h for h in heads_norm)
        has_jock  = any(("騎手" in h) or ("騎手名" in h) for h in heads_norm)
        if has_horse and has_jock:
            return t, heads_raw, {h:i for i,h in enumerate(heads_norm)}
    return None, None, None

def extract_basic_meta(text_all: str):
    m_date = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text_all)
    ymd = f"{int(m_date.group(1)):04d}{int(m_date.group(2)):02d}{int(m_date.group(3)):02d}" if m_date else datetime.now().strftime("%Y%m%d")
    m_place = re.search(r"\d+\s*回\s*(札幌|函館|福島|新潟|東京|中山|中京|京都|阪神|小倉)\s*\d+\s*日", text_all)
    place = m_place.group(1) if m_place else "不明"
    m_r1 = re.search(r"(\d{1,2})\s*レース", text_all)
    m_r2 = re.search(r"(\d{1,2})\s*R", text_all)
    race_no = f"{int((m_r1 or m_r2).group(1))}R" if (m_r1 or m_r2) else "R"
    return ymd, place, race_no

# ================== URL探索（複数戦略） ==================
def try_fetch(url: str, debug_log=None):
    """URLからHTMLを取得し、出馬表として有効か判定"""
    try:
        r = requests.get(url, headers={
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }, timeout=8)
        if debug_log:
            debug_log(f"試行: {url[:90]}... → {r.status_code}")
    except Exception as e:
        if debug_log:
            debug_log(f"接続失敗: {str(e)[:50]}")
        return None
    
    if r.status_code != 200:
        return None
    
    r.encoding = r.apparent_encoding
    soup = BeautifulSoup(r.text, "lxml")
    
    # 出馬表ページの特徴をチェック
    table, _, _ = find_table_and_headers(soup)
    if not table:
        return None
    
    # レース名要素の存在確認
    race_name = soup.select_one(".race_name") or soup.find(string=re.compile(r"\d+回.*(札幌|函館|福島|新潟|東京|中山|中京|京都|阪神|小倉)"))
    if not race_name:
        return None
    
    if debug_log:
        debug_log(f"  ✓ 有効な出馬表ページを検出！")
    return soup

def strategy_1_pattern_analysis(yyyymmdd: str, place: str, race_no: int, debug_log=None):
    """
    戦略1: 提供されたURLパターンの分析
    pw01dde01 05 20250501 11 20251108 /EB
            場所 開催日   R# 今日の日付
    
    サフィックスの規則性を探る
    """
    code = VENUE_CODE.get(place)
    if not code:
        return None, None
    
    if debug_log:
        debug_log("【戦略1】パターン分析による推測")
    
    # 今日の日付とその前後
    today = datetime.now()
    access_dates = [
        today.strftime("%Y%m%d"),
        (today + timedelta(days=1)).strftime("%Y%m%d"),
        (today - timedelta(days=1)).strftime("%Y%m%d"),
    ]
    
    # 観測されたサフィックスとその周辺
    # EB=235, 39=57, 1B=27 (16進数→10進数)
    # これらから規則性を推測
    suffixes = ["EB", "39", "1B", "E9", "EA", "EC", "37", "38", "3A", "19", "1A", "1C", "1D"]
    
    race_variants = [f"{race_no:02d}", f"{race_no}"]
    endpoints = ["accessD.html", "accessS.html"]
    
    tried = 0
    for endpoint in endpoints:
        for rn in race_variants:
            for access_date in access_dates:
                for suffix in suffixes:
                    cname = f"pw01dde{code}{yyyymmdd}{rn}{access_date}/{suffix}"
                    url = f"https://www.jra.go.jp/JRADB/{endpoint}?CNAME={cname}"
                    tried += 1
                    
                    soup = try_fetch(url, debug_log)
                    if soup:
                        return url, soup
                    
                    if tried > 50:
                        return None, None
    
    return None, None

def strategy_2_date_variations(yyyymmdd: str, place: str, race_no: int, debug_log=None):
    """
    戦略2: 日付バリエーション探索
    開催日と回数の組み合わせパターン
    """
    code = VENUE_CODE.get(place)
    if not code:
        return None, None
    
    if debug_log:
        debug_log("【戦略2】日付バリエーション探索")
    
    # 開催日の前後も試す（週末開催などの可能性）
    dt = datetime.strptime(yyyymmdd, "%Y%m%d")
    race_dates = [
        dt.strftime("%Y%m%d"),
        (dt - timedelta(days=1)).strftime("%Y%m%d"),
        (dt + timedelta(days=1)).strftime("%Y%m%d"),
    ]
    
    today = datetime.now().strftime("%Y%m%d")
    race_variants = [f"{race_no:02d}", f"{race_no}"]
    
    # より広範囲のサフィックス
    for i in range(256):
        suffix = f"{i:02X}"
        for race_date in race_dates:
            for rn in race_variants:
                cname = f"pw01dde{code}{race_date}{rn}{today}/{suffix}"
                url = f"https://www.jra.go.jp/JRADB/accessD.html?CNAME={cname}"
                
                soup = try_fetch(url, debug_log)
                if soup:
                    return url, soup
                
                # 20件試して見つからなければ次の日付へ
                if i > 20:
                    break
    
    return None, None

def strategy_3_scrape_pages(yyyymmdd: str, place: str, race_no: int, debug_log=None):
    """
    戦略3: JRAページから全リンクを抽出
    """
    if debug_log:
        debug_log("【戦略3】JRAページからリンク抽出")
    
    urls_to_scrape = [
        f"https://www.jra.go.jp/keiba/thisweek/",
        f"https://www.jra.go.jp/keiba/thisweek/{yyyymmdd}/",
        f"https://www.jra.go.jp/",
    ]
    
    all_links = set()
    
    for page_url in urls_to_scrape:
        try:
            if debug_log:
                debug_log(f"ページ取得: {page_url}")
            
            r = requests.get(page_url, headers={
                "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }, timeout=10)
            
            if r.status_code != 200:
                continue
            
            r.encoding = r.apparent_encoding
            soup = BeautifulSoup(r.text, "lxml")
            
            # すべてのリンクとonClickを抽出
            for elem in soup.find_all(["a", "button", "div", "span"]):
                # href属性
                href = elem.get("href", "")
                if "JRADB" in href and "CNAME" in href:
                    all_links.add(href if href.startswith("http") else f"https://www.jra.go.jp{href}")
                
                # onClick属性からCNAME抽出
                onclick = elem.get("onclick", "")
                match = re.search(r"CNAME=([^'\"&\s]+)", onclick)
                if match:
                    cname = match.group(1)
                    url = f"https://www.jra.go.jp/JRADB/accessD.html?CNAME={cname}"
                    all_links.add(url)
            
            if debug_log:
                debug_log(f"  {len(all_links)}個のJRADBリンクを発見")
        
        except Exception as e:
            if debug_log:
                debug_log(f"ページ取得エラー: {e}")
    
    # 抽出したリンクを試行
    code = VENUE_CODE.get(place)
    for url in all_links:
        # 場所コードとレース番号を含むURLを優先
        if code in url or str(race_no) in url or f"{race_no:02d}" in url:
            if debug_log:
                debug_log(f"候補リンクを検証: {url[:80]}...")
            soup = try_fetch(url, debug_log)
            if soup:
                # 本当に該当レースか確認
                text = soup.get_text()
                if place in text and (f"{race_no}R" in text or f"{race_no}レース" in text):
                    return url, soup
    
    return None, None

def build_jra_url_and_soup(yyyymmdd: str, place: str, race_no: int, status_cb=None, debug_log=None):
    """複数の戦略を順次試行"""
    
    strategies = [
        ("パターン分析", strategy_1_pattern_analysis),
        ("日付バリエーション", strategy_2_date_variations),
        ("ページスクレイピング", strategy_3_scrape_pages),
    ]
    
    for name, strategy in strategies:
        if status_cb:
            status_cb(f"{name}で探索中...")
        if debug_log:
            debug_log(f"\n{'='*60}")
            debug_log(f"戦略: {name}")
            debug_log(f"{'='*60}")
        
        try:
            url, soup = strategy(yyyymmdd, place, race_no, debug_log)
            if url and soup:
                if debug_log:
                    debug_log(f"\n✓ 成功！ URL発見: {url}")
                return url, soup
        except Exception as e:
            if debug_log:
                debug_log(f"戦略失敗: {e}")
    
    return None, None

# ================== 抽出＆Excel ==================
def fetch_rows_and_meta(url: str, soup: BeautifulSoup | None = None):
    if soup is None:
        r = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=15)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        soup = BeautifulSoup(r.text, "lxml")

    table, _, header_map = find_table_and_headers(soup)
    if not table:
        raise RuntimeError("出馬表テーブル（『馬名』『騎手』ヘッダー）が見つかりません。")

    col_umaban = find_col_index(header_map, ["馬番","馬番号"])
    col_horse  = find_col_index(header_map, ["馬名"])
    col_jock   = find_col_index(header_map, ["騎手","騎手名"])
    if col_horse is None or col_jock is None:
        raise RuntimeError("『馬名』『騎手(騎手名)』列を特定できませんでした。")

    trs = table.find_all("tr")
    start_idx = 1 if trs and trs[0].find_all("th") else 0

    rows = []
    for tr in trs[start_idx:]:
        tds = tr.find_all(["td","th"])
        if len(tds) <= max(col_horse, col_jock):
            continue

        horse  = clean_name(anchor_text(tds[col_horse]))
        jockey = clean_name(anchor_text(tds[col_jock]))
        if not horse or re.fullmatch(r"\d+", horse):
            continue
        if not jockey or jockey == "-":
            continue

        umaban = ""
        if col_umaban is not None and len(tds) > col_umaban:
            m = re.search(r"\d{1,2}", anchor_text(tds[col_umaban]).strip())
            umaban = m.group(0) if m else ""
        else:
            m = re.match(r"\D*(\d{1,2})\D*", anchor_text(tds[0]) if tds else "")
            umaban = m.group(1) if m else ""

        rows.append((umaban, horse, jockey))

    if not rows:
        raise RuntimeError("馬番／馬名／騎手名の抽出結果が空でした。")

    race_el = soup.select_one(".race_name")
    race_title = race_el.get_text(strip=True).split("|")[0].strip() if race_el else ""

    text_all = soup.get_text(" ", strip=True)
    ymd, place, race_no = extract_basic_meta(text_all)
    if not race_title:
        race_title = f"{place}{race_no}"

    filename = f"{ymd}_{place}_{race_no}.xlsx"
    return rows, filename, race_title, url

def save_to_desktop(rows, filename, race_title):
    desktop = os.path.join(os.path.expanduser("~"), "デスクトップ")
    os.makedirs(desktop, exist_ok=True)
    path = os.path.join(desktop, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "出馬表"

    ws.merge_cells("A1:E1")
    t = ws["A1"]; t.value = race_title
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.font = Font(bold=True, size=18)
    t.fill = PatternFill(start_color="FADADD", end_color="FADADD", fill_type="solid")
    ws.row_dimensions[1].height = 30

    ws.append(["馬番","馬名","騎手名","評価","短評"])
    for umaban, horse, jockey in rows:
        ws.append([umaban, horse, jockey, "", ""])

    light_blue = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    bold = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for c in range(1, 6):
        cell = ws.cell(row=2, column=c)
        cell.fill = light_blue
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws.max_row + 1):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50

    wb.save(path)
    return path

# ================== GUI ==================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("JRA出馬表 取り込みツール（3戦略併用版） - v0.14")
        self.geometry("750x550"); self.resizable(True, True)

        frm = ttk.Frame(self, padding=12); frm.pack(fill="both", expand=True)

        row = 0
        ttk.Label(frm, text="取得方法").grid(row=row, column=0, sticky="w")
        self.mode = tk.StringVar(value="select")
        ttk.Radiobutton(frm, text="日付・場所・レースを選択", variable=self.mode, value="select").grid(row=row, column=1, sticky="w")
        ttk.Radiobutton(frm, text="URLを直接入力",     variable=self.mode, value="url").grid(row=row, column=2, sticky="w")

        row += 1
        ttk.Label(frm, text="日付").grid(row=row, column=0, sticky="e", pady=6)
        if TKCAL_OK:
            self.date_widget = DateEntry(frm, width=14, date_pattern="yyyyMMdd")
        else:
            self.date_widget = ttk.Entry(frm, width=16)
            self.date_widget.insert(0, datetime.now().strftime("%Y%m%d"))
        self.date_widget.grid(row=row, column=1, sticky="w", pady=6)
        self.cal_btn = ttk.Button(frm, text="📅", width=3, command=self.open_calendar); self.cal_btn.grid(row=row, column=2, sticky="w")

        ttk.Label(frm, text="場所").grid(row=row, column=3, sticky="e")
        self.cmb_place = ttk.Combobox(frm, values=VENUES, width=8, state="readonly"); self.cmb_place.grid(row=row, column=4, sticky="w"); self.cmb_place.set("京都")

        row += 1
        ttk.Label(frm, text="レース").grid(row=row, column=0, sticky="e")
        self.cmb_race = ttk.Combobox(frm, values=[f"{i}R" for i in range(1,13)], width=6, state="readonly"); self.cmb_race.grid(row=row, column=1, sticky="w"); self.cmb_race.set("11R")

        row += 1
        ttk.Label(frm, text="URL（任意：直入力する場合）").grid(row=row, column=0, sticky="e")
        self.ent_url = ttk.Entry(frm, width=56); self.ent_url.grid(row=row, column=1, columnspan=4, sticky="ew")

        row += 1
        self.btn = ttk.Button(frm, text="Excelを作成（デスクトップに保存）", command=self.run_fetch); self.btn.grid(row=row, column=0, columnspan=5, sticky="ew", pady=12)

        row += 1
        self.status = ttk.Label(frm, text=f"待機中 / 方式: 3戦略併用 / Calendar: {'ON' if TKCAL_OK else 'OFF'}"); self.status.grid(row=row, column=0, columnspan=5, sticky="w")

        row += 1
        ttk.Label(frm, text="デバッグログ (URLが見つかったら、次回用にメモしてください):").grid(row=row, column=0, columnspan=5, sticky="w", pady=(10,0))
        
        row += 1
        self.debug_text = scrolledtext.ScrolledText(frm, height=12, width=80, state="disabled")
        self.debug_text.grid(row=row, column=0, columnspan=5, sticky="nsew", pady=5)

        frm.rowconfigure(row, weight=1)
        for i in range(5): frm.columnconfigure(i, weight=1)

    def debug_log(self, msg):
        """デバッグログに追記"""
        self.after(0, lambda: self._append_debug(msg))
    
    def _append_debug(self, msg):
        self.debug_text.config(state="normal")
        self.debug_text.insert(tk.END, msg + "\n")
        self.debug_text.see(tk.END)
        self.debug_text.config(state="disabled")

    def open_calendar(self):
        if not TKCAL_OK:
            top = tk.Toplevel(self); top.title("日付を入力（YYYYMMDD）"); top.resizable(False, False)
            ttk.Label(top, text="tkcalendar が見つかりません。\nYYYYMMDD 形式で入力してください。").pack(padx=10, pady=10)
            ent = ttk.Entry(top); ent.insert(0, self._current_ymd_or_today()); ent.pack(padx=10, pady=6)
            ttk.Button(top, text="この日付で決定", command=lambda: (self.set_date_value(ent.get().strip()), top.destroy())).pack(pady=8)
            return
        cur = self._current_ymd_or_today(); y, m, d = int(cur[:4]), int(cur[4:6]), int(cur[6:8])
        top = tk.Toplevel(self); top.title("日付を選択"); top.resizable(False, False)
        cal = Calendar(top, year=y, month=m, day=d, date_pattern="yyyy-mm-dd"); cal.pack(padx=10, pady=10)
        ttk.Button(top, text="この日付で決定", command=lambda: (self.set_date_value(cal.get_date().replace("-", "")), top.destroy())).pack(pady=8)

    def _current_ymd_or_today(self) -> str:
        try:
            cur = self.date_widget.get().strip()
            if re.fullmatch(r"\d{8}", cur): return cur
        except Exception:
            pass
        return datetime.now().strftime("%Y%m%d")

    def set_date_value(self, ymd: str):
        if re.fullmatch(r"\d{8}", ymd):
            if hasattr(self.date_widget, "delete"): self.date_widget.delete(0, tk.END)
            if hasattr(self.date_widget, "insert"): self.date_widget.insert(0, ymd)
        else:
            messagebox.showwarning("日付形式エラー", "YYYYMMDD で入力してください。")

    def run_fetch(self):
        # デバッグログをクリア
        self.debug_text.config(state="normal")
        self.debug_text.delete(1.0, tk.END)
        self.debug_text.config(state="disabled")
        
        mode = self.mode.get()
        url_manual = self.ent_url.get().strip()

        if mode == "url" and url_manual:
            self._start_job(url_manual, soup=None)
            return

        ymd = self._current_ymd_or_today()
        place = self.cmb_place.get().strip()
        race = self.cmb_race.get().strip()

        if not (re.fullmatch(r"\d{8}", ymd) and place in VENUES and re.fullmatch(r"\d{1,2}R", race)):
            messagebox.showwarning("入力値エラー", "日付はカレンダーで選択、場所はリストから、レースは1R〜12Rを選択してください。")
            return

        rno = int(race[:-1])
        self.btn.config(state="disabled"); self.status.config(text="3つの戦略で探索中...")
        self.debug_log(f"探索開始: {ymd} {place} {rno}R")
        self.debug_log("戦略1→戦略2→戦略3の順で試行します\n")
        threading.Thread(target=self._auto_and_fetch, args=(ymd, place, rno), daemon=True).start()

    def _auto_and_fetch(self, ymd, place, rno):
        try:
            url, soup = build_jra_url_and_soup(
                ymd, place, rno, 
                status_cb=lambda s: self.after(0, lambda: self.status.config(text=s)),
                debug_log=self.debug_log
            )
            if not url:
                self._done("該当するレースのURLが見つかりませんでした。\n\n【推奨】JRA公式サイトで該当レースのURLをコピーし、\n上部の「URL（任意）」欄に貼り付けて実行してください。")
                return
            rows, filename, race_title, used_url = fetch_rows_and_meta(url, soup)
            out = save_to_desktop(rows, filename, race_title)
            self._done(f"保存完了：{out}\n\n使用URL（次回用にメモ推奨）：\n{used_url}")
        except Exception as e:
            self._done(f"エラー：{e}")

    def _start_job(self, url, soup=None):
        self.btn.config(state="disabled"); self.status.config(text="取得中…")
        self.debug_log(f"URL直接取得: {url}")
        threading.Thread(target=self._do_fetch, args=(url, soup), daemon=True).start()

    def _do_fetch(self, url, soup):
        try:
            rows, filename, race_title, used_url = fetch_rows_and_meta(url, soup)
            out = save_to_desktop(rows, filename, race_title)
            self._done(f"保存完了：{out}\n使用URL：{used_url}")
        except Exception as e:
            self._done(f"エラー：{e}")

    def _done(self, msg):
        self.after(0, lambda: self._finish_ui(msg))

    def _finish_ui(self, msg):
        self.btn.config(state="normal"); self.status.config(text=msg.splitlines()[0])
        self.debug_log(f"\n{'='*60}\n{msg}\n{'='*60}")
        (messagebox.showinfo if msg.startswith("保存完了") else messagebox.showwarning)("結果", msg)

# ---------------- main ----------------
if __name__ == "__main__":
    App().mainloop()